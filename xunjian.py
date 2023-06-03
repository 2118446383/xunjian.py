from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout,
                             QPushButton, QFileDialog, QWidget, QLabel, QSpacerItem,
                             QSizePolicy, QTextEdit, QDialog, QGridLayout,
                             QComboBox, QMessageBox, QTableWidget, QTableWidgetItem, QDialogButtonBox, QListWidget, QAbstractItemView, QListWidgetItem)
from PyQt6.QtCore import QThread, Qt
from PyQt6.QtGui import QCloseEvent, QIcon
from PyQt6 import sip
from PyQt6.QtCore import QObject, pyqtSlot, pyqtSignal
from openpyxl import load_workbook
import pickle
import sys
import io

from main import main, IT
from ip_scanner import ITScanner

class CustomStream(QObject):
    text_written = pyqtSignal(str)  # 定义一个新信号以传递写入的文本

    class _TextIO(io.TextIOWrapper):
        def __init__(self, custom_stream):
            super().__init__(io.StringIO(), encoding="utf-8", newline="\r\n")
            self.custom_stream = custom_stream

        def write(self, text):
            self.custom_stream.text_written.emit(text)


    def __init__(self, parent=None):
        super().__init__(parent)
        self._text_io = self._TextIO(self)

    def write(self, text):
        self._text_io.write(text)

    def flush(self):
        pass

class Worker(QThread):
    import_finished = pyqtSignal(str)
    inspection_finished = pyqtSignal()

    def __init__(self, operation_type, file_path=None):
        super().__init__()
        self.operation_type = operation_type
        self.file_path = file_path

    def run(self):
        it_instance = IT.get_instance()

        if self.operation_type == "import_hosts*":
            main(import_hosts=self.file_path, is_gui_mode=True)
            self.import_finished.emit(self.file_path)

        elif self.operation_type == "run_main":
            if len(it_instance.hosts) == 0:
                print("尚未导入任何主机，请先导入主机文件")
            else:
                main(run=True, is_gui_mode=True)
                self.inspection_finished.emit()

class IpScannerThread(QThread):
    def __init__(self, file_path=None):
        super().__init__()
        self.file_path = file_path

    def run(self):
        scanner = ITScanner()  # 创建一个新的ITScanner实例
        scanner.read_targets_from_xlsx(self.file_path)
        scanner.run_port_scanner()

class PortScannerDialog(QDialog):
    def __init__(self, parent=None, worker_threads=None):
        super().__init__(parent)
        self.setWindowTitle("扫描ip开放端口")
        self.setFixedSize(400, 200)

        if not worker_threads:
            worker_threads = []
        self.worker_threads = worker_threads

        layout = QVBoxLayout(self)

        self.select_file_button = QPushButton("选择文件")
        self.select_file_button.clicked.connect(self.select_file)
        layout.addWidget(self.select_file_button)

        self.modify_file_button = QPushButton("修改文件")
        self.modify_file_button.clicked.connect(self.modify_file)
        self.modify_file_button.setEnabled(False)
        layout.addWidget(self.modify_file_button)

        self.start_scanner_button = QPushButton("开始执行")
        self.start_scanner_button.clicked.connect(self.start_scanner)
        self.start_scanner_button.setEnabled(False)
        layout.addWidget(self.start_scanner_button)

    def select_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择渗透测试文件", "", "Excel 文件 (*.xlsx)")
        if file:
            self.selected_file = file
            self.start_scanner_button.setEnabled(True)
            self.modify_file_button.setEnabled(True)

    def modify_file(self):
        modify_dialog = ModifyFileDialog(self.selected_file, self)
        modify_dialog.exec()

    def start_scanner(self):
        worker = IpScannerThread(self.selected_file)
        worker.finished.connect(worker.deleteLater)
        worker.start()
        self.worker_threads.append(worker)
        self.close()

class ModifyFileDialog(QDialog):
    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("修改文件")
        self.setFixedSize(800, 400)

        layout = QVBoxLayout(self)

        self.command_table = QTableWidget()
        layout.addWidget(self.command_table)

        save_button = QPushButton("保存修改")
        save_button.clicked.connect(self.save_changes)
        layout.addWidget(save_button)

        add_row_button = QPushButton("添加行")
        add_row_button.clicked.connect(self.add_row)
        layout.addWidget(add_row_button)

        delete_row_button = QPushButton("删除行")
        delete_row_button.clicked.connect(self.delete_row)
        layout.addWidget(delete_row_button)

        self.file_path = file_path
        self.load_file()

    def load_file(self):
        self.workbook = load_workbook(self.file_path)
        self.worksheet = self.workbook.active

        self.command_table.setRowCount(self.worksheet.max_row)
        self.command_table.setColumnCount(self.worksheet.max_column)

        for row_index, row in enumerate(self.worksheet.iter_rows()):
            for col_index, cell in enumerate(row):
                item = QTableWidgetItem(str(cell.value))
                self.command_table.setItem(row_index, col_index, item)

    def save_changes(self):
        # Clear the original worksheet
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.value = None

        # Write the modified data back to the original worksheet
        for row_index in range(self.command_table.rowCount()):
            for col_index in range(self.command_table.columnCount()):
                item = self.command_table.item(row_index, col_index)
                if item is not None:
                    value = item.text()
                    self.worksheet.cell(row=row_index + 1, column=col_index + 1, value=value)

        self.workbook.save(self.file_path)
        QMessageBox.information(self, "成功", "已成功保存更改！")

    def add_row(self):
        row_position = self.command_table.rowCount()
        self.command_table.insertRow(row_position)

    def delete_row(self):
        if self.command_table.currentRow() >= 0:
            row_position = self.command_table.currentRow()
            self.command_table.removeRow(row_position)
        else:
            QMessageBox.warning(self, "警告", "请选择要删除的行。")

class HostSelectionDialog(QDialog):
    def __init__(self, hosts, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择要巡检的主机")
        self.setFixedSize(400, 300)

        layout = QVBoxLayout(self)

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        for host in hosts:
            item = QListWidgetItem(f"{host['ip']}")  # 修改这一行，移除 {host['hostname']}
            item.setData(Qt.ItemDataRole.UserRole, host['ip'])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.list_widget.addItem(item)
        layout.addWidget(self.list_widget)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_selected_ips(self):
        selected_ips = []
        for index in range(self.list_widget.count()):
            item = self.list_widget.item(index)
            if item.checkState() == Qt.CheckState.Checked:
                ip = item.data(Qt.ItemDataRole.UserRole)
                selected_ips.append(ip)
        return selected_ips

class CommandDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("修改巡检命令")
        self.setFixedSize(800, 400)

        layout = QGridLayout(self)

        hint_text = "温馨提示："
        self.hint_label = QLabel(hint_text)
        font = self.hint_label.font()
        font.setPointSize(10)
        self.hint_label.setFont(font)
        self.hint_label.setStyleSheet("color: red")
        layout.addWidget(self.hint_label, 4, 0, 1, 2)  # 将标签添加到布局中

        warning_text = "配置文件.txt，设备cpu负载.txt，设备内存负载.txt，设备温度和电源和风扇运转参数.txt，系统版本信息设备运行时间.txt 这些为默认名称，不可随意修改，关联巡检报告是否正常生成"
        self.warning_label = QLabel(warning_text)
        self.warning_label.setWordWrap(True)
        layout.addWidget(self.warning_label, 4, 0, 3, 2)  # 将标签添加到布局中

        self.company_combo_box = QComboBox()
        self.company_combo_box.addItem("Cisco or Ruijie", "cisco")
        self.company_combo_box.addItem("H3C or Huawei", "h3c")
        layout.addWidget(self.company_combo_box, 0, 0)

        self.load_commands_button = QPushButton("加载命令")
        self.load_commands_button.clicked.connect(self.load_commands)
        layout.addWidget(self.load_commands_button, 0, 1)

        self.command_table = QTableWidget(0, 3)
        self.command_table.setHorizontalHeaderLabels(["文本文件名", "巡检命令", "延迟时间"])
        self.command_table.horizontalHeader().setStretchLastSection(True)  # 最后一列填充满表格
        self.command_table.setEditTriggers(QTableWidget.EditTrigger.AllEditTriggers)
        layout.addWidget(self.command_table, 1, 0, 1, 2)

        self.add_command_button = QPushButton("添加")
        self.add_command_button.clicked.connect(self.add_command)
        layout.addWidget(self.add_command_button, 2, 0)

        self.delete_command_button = QPushButton("删除")
        self.delete_command_button.clicked.connect(self.delete_command)
        layout.addWidget(self.delete_command_button, 2, 1)

        save_button = QPushButton("保存")
        save_button.clicked.connect(self.save_changes)
        layout.addWidget(save_button, 3, 0, 1, 2)  # 这一行代码应该在 __init__ 方法中

    def add_command(self):
        row_count = self.command_table.rowCount()
        self.command_table.insertRow(row_count)  # 在表格末尾添加新行

    def delete_command(self):
        selected_rows = sorted(set(item.row() for item in self.command_table.selectedItems()))

        # 删除多个选定行（从最后一个开始）
        for row in reversed(selected_rows):
            self.command_table.removeRow(row)

    def load_commands(self):
        company = self.company_combo_box.currentData()
        it_instance = IT.get_instance()

        commands = []  # 添加默认值（空列表）

        if company == "cisco":
            commands = it_instance.cisco_commands
        elif company in ["h3c", "huawei"]:
            commands = it_instance.h3c_huawei_commands

        self.command_table.setRowCount(len(commands))
        for row, (filename, cmd_list, delay) in enumerate(commands):
            self.command_table.setItem(row, 0, QTableWidgetItem(filename))
            self.command_table.setItem(row, 1, QTableWidgetItem(", ".join(cmd_list)))
            self.command_table.setItem(row, 2, QTableWidgetItem(str(delay)))

    def save_changes(self):
        company = self.company_combo_box.currentData()
        it_instance = IT.get_instance()

        commands = []
        for row in range(self.command_table.rowCount()):
            filename = self.command_table.item(row, 0).text().strip()
            command = self.command_table.item(row, 1).text().strip()
            delay = self.command_table.item(row, 2).text().strip()

            if not filename or not command or not delay:
                QMessageBox.warning(self, "错误", "所有字段都不能为空。")
                return

            try:
                delay = int(delay)
            except ValueError:
                QMessageBox.warning(self, "错误", "延迟时间必须是整数。")
                return

            commands.append((filename, command.split(", "), delay))

        it_instance.update_commands(company, commands)

        # 将命令列表保存到文件中
        with open(it_instance.commands_file, 'wb') as f:
            pickle.dump((it_instance.cisco_commands, it_instance.h3c_huawei_commands), f)

        self.close()


class NetworkInspectionApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.worker_threads = []

        self.setWindowTitle("网络设备巡检工具")
        self.setFixedSize(1000, 600)

        # 设置窗口图标
        self.setWindowIcon(QIcon("logo.png"))  # 将 "your_icon.png" 替换为图标文件的名称

        layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        layout.addLayout(left_layout)

        title_label = QLabel("网络设备巡检工具")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; color: white;")
        left_layout.addWidget(title_label)

        # 修改窗口的背景颜色
        self.setStyleSheet("""
            QMainWindow {
                background-color: #324259;
            }
        """)

        left_layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        button_style = """
            QPushButton {
                background-color: qlineargradient(spread:pad, x1:0, y1:0.5, x2:1, y2:0.5, stop:0 rgba(50, 66, 89, 255), stop:1 rgba(3, 169, 244, 255));
                color: white;
                border-radius: 5px;
                font-size: 18px;
                padding: 10px 30px;
            }

            QPushButton:hover {
                background-color: qlineargradient(spread:pad, x1:0, y1:0.5, x2:1, y2:0.5, stop:0 rgba(50, 66, 89, 255), stop:1 rgba(63, 81, 181, 255));
            }
        """

        load_hosts_button = QPushButton("导入主机")
        load_hosts_button.clicked.connect(self.load_hosts)
        load_hosts_button.setFixedHeight(50)
        load_hosts_button.setStyleSheet(button_style)
        left_layout.addWidget(load_hosts_button)

        self.select_inspection_hosts_button = QPushButton("选择巡检主机")
        self.select_inspection_hosts_button.clicked.connect(self.select_inspection_hosts)
        self.select_inspection_hosts_button.setFixedHeight(50)
        self.select_inspection_hosts_button.setStyleSheet(button_style)
        self.select_inspection_hosts_button.setEnabled(False)  # 默认禁用按钮
        left_layout.addWidget(self.select_inspection_hosts_button)

        modify_command_button = QPushButton("修改巡检命令")
        modify_command_button.clicked.connect(self.modify_inspection_command)
        modify_command_button.setFixedHeight(50)
        modify_command_button.setStyleSheet(button_style)
        left_layout.addWidget(modify_command_button)

        run_button = QPushButton("开始巡检")
        run_button.clicked.connect(self.run_inspection)
        run_button.setFixedHeight(50)
        run_button.setStyleSheet(button_style)
        left_layout.addWidget(run_button)

        run_port_scanner_button = QPushButton("扫描ip开放端口")
        run_port_scanner_button.clicked.connect(self.run_port_scanner)
        run_port_scanner_button.setFixedHeight(50)
        run_port_scanner_button.setStyleSheet(button_style)
        left_layout.addWidget(run_port_scanner_button)

        close_button = QPushButton("关闭程序")
        close_button.clicked.connect(self.close)
        close_button.setFixedHeight(50)
        close_button.setStyleSheet(button_style)
        left_layout.addWidget(close_button)

        left_layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        self.output_text_edit = QTextEdit()
        self.output_text_edit.setReadOnly(True)
        self.output_text_edit.setStyleSheet("background-color: #f0f0f0;") # 修改输出文本框的背景颜色
        layout.addWidget(self.output_text_edit)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        sys.stdout = CustomStream(self)
        sys.stdout.text_written.connect(self.update_output)

    def run_port_scanner(self):
        port_scanner_dialog = PortScannerDialog(self, self.worker_threads)
        port_scanner_dialog.exec()

    @pyqtSlot(str)
    def update_output(self, text):
        self.output_text_edit.append(text.strip())
        scrollbar = self.output_text_edit.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def load_hosts(self):
        hosts_file, _ = QFileDialog.getOpenFileName(self, "选择主机文件", "", "Hosts Excel 文件 (hosts*.xlsx)")

        if hosts_file:
            worker = Worker("import_hosts*", hosts_file)
            worker.import_finished.connect(self.on_import_finished)
            worker.finished.connect(worker.deleteLater)
            worker.start()
            self.worker_threads.append(worker)

    def select_inspection_hosts(self):
        it_instance = IT.get_instance()
        host_selection_dialog = HostSelectionDialog(it_instance.hosts, self)
        result = host_selection_dialog.exec()

        if result == QDialog.DialogCode.Accepted:
            selected_ips = host_selection_dialog.get_selected_ips()
            it_instance.update_hosts(selected_ips)  # 确保调用了 update_hosts 方法
            print(f"选择了以下主机进行巡检: {', '.join(selected_ips)}")

    def run_inspection(self):
        worker = Worker("run_main")
        worker.inspection_finished.connect(self.on_inspection_finished)
        worker.finished.connect(worker.deleteLater)
        worker.start()
        self.worker_threads.append(worker)

    def modify_inspection_command(self):
        command_dialog = CommandDialog(self)
        command_dialog.exec()

    def on_import_finished(self, hosts_file):
        print(f"成功导入主机: {hosts_file}")
        self.select_inspection_hosts_button.setEnabled(True)  # 启用选择巡检主机按钮

    def on_inspection_finished(self):
        print("巡检任务已完成")

    def wait_for_workers(self):
        for worker in self.worker_threads:
            if not sip.isdeleted(worker) and worker.isRunning():
                worker.wait()

    def closeEvent(self, event: QCloseEvent):
        for worker in self.worker_threads:
            if not sip.isdeleted(worker) and worker.isRunning():
                worker.wait()

        event.accept()

if __name__ == "__main__":
    app = QApplication([])
    window = NetworkInspectionApp()
    window.show()
    app.exec()
    window.wait_for_workers()